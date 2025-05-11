from django.http import HttpResponse
from .models import Saving
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

def convert_saving_log_to_excel(request, slug) :
    saving = Saving.objects.get(slug = slug) 
    logs = saving.saving_entries.all()
    if logs : 
        wb = Workbook()
        ws = wb.active
        ws.title = f"Catatan Tabungan {saving.title}"

        # Header row styling
        header = ['No', 'Nominal', 'Total', 'Description', 'Status', 'Datetime']
        ws.append(header)

        # Apply header styles
        for cell in ws[1]:
            cell.font = Font(bold=True)  # Bold font for header
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Center alignment

        # Color fill for "masuk" (green) and "keluar" (red)
        masuk_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # light green
        keluar_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # light red

        # Loop through the logs and append rows
        for idx, log in enumerate(logs, start=1):
            # Determine fill based on status
            if log.status == 'masuk':
                row_fill = masuk_fill
            elif log.status == 'keluar':
                row_fill = keluar_fill
            else:
                row_fill = None  # No fill if status is something else

            # Append the row
            ws.append([
                idx,
                float(log.nominal) if log.nominal else 0.00,
                float(log.total),
                log.description,
                log.status,
                log.datetime.strftime('%Y-%m-%d %H:%M:%S'),
            ])

            # Apply color to the entire row based on the status
            for cell in ws[idx + 1]:  # `idx + 1` because `ws.append` increases the row index
                if row_fill:
                    cell.fill = row_fill

        # Set column widths for a cleaner look
        for col in range(1, len(header) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Apply border to all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        # Adjust alignment
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add response headers for download
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Tabungan {saving.title}.xlsx'
        wb.save(response)
        return response
    else :
        print('dak pacak')
        return HttpResponse(status = 204)